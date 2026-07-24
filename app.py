import os
import io
import json
import time
import gzip
from flask import Flask, jsonify, send_from_directory, request, Response
from threading import Lock

from dashboard_logic import merge_and_build_tree, tree_payload, normalize_id, apply_authoritative_location_sales

app = Flask(__name__, static_folder="static", static_url_path="")

# --- Configuration ---
# Note: this server no longer needs any Redash credentials at all — those
# stay on whichever machine runs push_local.py. This server only needs a
# shared secret so it can tell a legitimate push apart from a random request.
PUSH_TOKEN = os.environ.get("PUSH_TOKEN", "")
DASHBOARD_TITLE = os.environ.get("DASHBOARD_TITLE", "ZM / RM / Location Performance")

# --- Password protection for the dashboard itself ---
# Uses a single shared username/password via the browser's built-in login
# prompt (HTTP Basic Auth) -- not a custom login page, but simple and secure
# over HTTPS (which Render provides by default).
#
# The two push endpoints (/api/push-data, /api/push-location-sales) are
# deliberately EXCLUDED from this check -- those are called by push_local.py,
# an automated script with no browser and no way to type a password. They
# already have their own protection via PUSH_TOKEN, checked separately in
# each of those routes below.
DASHBOARD_USERNAME = os.environ.get("DASHBOARD_USERNAME", "")
DASHBOARD_PASSWORD = os.environ.get("DASHBOARD_PASSWORD", "")
_PUSH_ONLY_PATHS = {"/api/push-data", "/api/push-location-sales"}


@app.before_request
def _require_dashboard_login():
    if request.path in _PUSH_ONLY_PATHS:
        return None
    if not DASHBOARD_USERNAME or not DASHBOARD_PASSWORD:
        return None  # not configured yet -- don't accidentally lock everyone out
    auth = request.authorization
    if not auth or auth.username != DASHBOARD_USERNAME or auth.password != DASHBOARD_PASSWORD:
        return Response(
            "Login required to view this dashboard.",
            401,
            {"WWW-Authenticate": 'Basic realm="ZM/RM/Location Dashboard"'},
        )
    return None


HIERARCHY_FILE = os.path.join(os.path.dirname(__file__), "hierarchy_upload.xlsx")
HIERARCHY_CACHE_FILE = os.path.join(os.path.dirname(__file__), "hierarchy_cache.json")
REDASH_CACHE_FILE = os.path.join(os.path.dirname(__file__), "redash_cache.json")
LOCATION_SALES_CACHE_FILE = os.path.join(os.path.dirname(__file__), "location_sales_cache.json")

_lock = Lock()
_state = {
    "redash_rows": None,        # list of raw dicts, received via /api/push-data
    "redash_pushed_at": 0,
    "location_sales_rows": None,   # list of raw dicts, received via /api/push-location-sales
    "location_sales_pushed_at": 0,
    "hierarchy_map": None,      # {employee_id: {am, rm, zm}}
    "hierarchy_uploaded_at": 0,
    "hierarchy_row_count": 0,
    "computed_by_vertical": {},  # cached merge_and_build_tree() results, keyed by vertical filter
}


# ---------------------------------------------------------------------------
# Disk persistence so a restart doesn't lose the last known-good data
# ---------------------------------------------------------------------------
def _load_disk_state():
    if os.path.exists(REDASH_CACHE_FILE):
        try:
            with open(REDASH_CACHE_FILE, "r") as f:
                payload = json.load(f)
                _state["redash_rows"] = payload.get("rows")
                _state["redash_pushed_at"] = payload.get("pushed_at", 0)
        except Exception:
            pass
    if os.path.exists(LOCATION_SALES_CACHE_FILE):
        try:
            with open(LOCATION_SALES_CACHE_FILE, "r") as f:
                payload = json.load(f)
                _state["location_sales_rows"] = payload.get("rows")
                _state["location_sales_pushed_at"] = payload.get("pushed_at", 0)
        except Exception:
            pass
    if os.path.exists(HIERARCHY_CACHE_FILE):
        try:
            with open(HIERARCHY_CACHE_FILE, "r") as f:
                payload = json.load(f)
                _state["hierarchy_map"] = payload.get("map")
                _state["hierarchy_uploaded_at"] = payload.get("uploaded_at", 0)
                _state["hierarchy_row_count"] = payload.get("row_count", 0)
        except Exception:
            pass


def _save_redash_cache():
    try:
        with open(REDASH_CACHE_FILE, "w") as f:
            json.dump({"rows": _state["redash_rows"], "pushed_at": _state["redash_pushed_at"]}, f)
    except Exception:
        pass


def _save_location_sales_cache():
    try:
        with open(LOCATION_SALES_CACHE_FILE, "w") as f:
            json.dump({"rows": _state["location_sales_rows"], "pushed_at": _state["location_sales_pushed_at"]}, f)
    except Exception:
        pass


def _save_hierarchy_cache():
    try:
        with open(HIERARCHY_CACHE_FILE, "w") as f:
            json.dump({
                "map": _state["hierarchy_map"],
                "uploaded_at": _state["hierarchy_uploaded_at"],
                "row_count": _state["hierarchy_row_count"],
            }, f)
    except Exception:
        pass


_load_disk_state()


# ---------------------------------------------------------------------------
# Receiving pushed Redash data (from push_local.py, run on someone's PC)
# ---------------------------------------------------------------------------
@app.route("/api/push-data", methods=["POST"])
def push_data():
    if not PUSH_TOKEN:
        return jsonify({"ok": False, "error": "Server has no PUSH_TOKEN configured"}), 500
    supplied = request.headers.get("X-Push-Token", "")
    if supplied != PUSH_TOKEN:
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    payload = request.get_json(force=True, silent=True) or {}
    rows = payload.get("rows")
    if not isinstance(rows, list):
        return jsonify({"ok": False, "error": "Expected JSON body like {\"rows\": [...]}"}), 400

    with _lock:
        _state["redash_rows"] = rows
        _state["redash_pushed_at"] = time.time()
        _state["computed_by_vertical"] = {}  # invalidate merged cache
        _save_redash_cache()

    return jsonify({"ok": True, "row_count": len(rows), "pushed_at": _state["redash_pushed_at"]})


@app.route("/api/push-location-sales", methods=["POST"])
def push_location_sales():
    """Receives query 13308's rows: authoritative per-location sales figures
    (WTD/MTD/M1), which replace the employee-summed numbers at the Location
    level (and their RM/ZM rollups) when viewing All Verticals."""
    if not PUSH_TOKEN:
        return jsonify({"ok": False, "error": "Server has no PUSH_TOKEN configured"}), 500
    supplied = request.headers.get("X-Push-Token", "")
    if supplied != PUSH_TOKEN:
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    payload = request.get_json(force=True, silent=True) or {}
    rows = payload.get("rows")
    if not isinstance(rows, list):
        return jsonify({"ok": False, "error": "Expected JSON body like {\"rows\": [...]}"}), 400

    with _lock:
        _state["location_sales_rows"] = rows
        _state["location_sales_pushed_at"] = time.time()
        _state["computed_by_vertical"] = {}  # invalidate merged cache
        _save_location_sales_cache()

    return jsonify({"ok": True, "row_count": len(rows), "pushed_at": _state["location_sales_pushed_at"]})


# ---------------------------------------------------------------------------
# Hierarchy upload / parse (unchanged — still uploaded directly to this server)
# ---------------------------------------------------------------------------
def parse_hierarchy_file(file_bytes):
    """Parse the weekly HC sheet. Column letters per the org's convention:
    B=Employee ID, H=L3 Name (AM), J=L4 Name (RM), L=L5 Name (ZM). Only L1s matter.
    Uses openpyxl directly (pure Python, no compiled deps)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    header = [str(h).strip() if h is not None else "" for h in header]

    required = ["Employee ID", "L3 Name", "L4 Name", "L5 Name", "Level"]
    missing = [c for c in required if c not in header]
    if missing:
        raise ValueError(f"Uploaded sheet is missing expected column(s): {missing}")

    col_idx = {name: header.index(name) for name in required}

    hierarchy_map = {}
    for row in rows_iter:
        if row is None:
            continue
        level = row[col_idx["Level"]]
        if level != "L1":
            continue
        eid = row[col_idx["Employee ID"]]
        if eid is None:
            continue
        hierarchy_map[normalize_id(eid)] = {
            "am": row[col_idx["L3 Name"]],
            "rm": row[col_idx["L4 Name"]],
            "zm": row[col_idx["L5 Name"]],
        }
    return hierarchy_map


@app.route("/api/upload-hierarchy", methods=["POST"])
def upload_hierarchy():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded"}), 400
    f = request.files["file"]
    file_bytes = f.read()
    try:
        hierarchy_map = parse_hierarchy_file(file_bytes)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Couldn't parse file: {e}"}), 400

    with _lock:
        _state["hierarchy_map"] = hierarchy_map
        _state["hierarchy_uploaded_at"] = time.time()
        _state["hierarchy_row_count"] = len(hierarchy_map)
        _state["computed_by_vertical"] = {}  # invalidate merged cache
        _save_hierarchy_cache()
        try:
            with open(HIERARCHY_FILE, "wb") as out:
                out.write(file_bytes)
        except Exception:
            pass

    return jsonify({"ok": True, "row_count": len(hierarchy_map)})


@app.route("/api/hierarchy-status")
def hierarchy_status():
    return jsonify({
        "ok": True,
        "loaded": _state["hierarchy_map"] is not None,
        "uploaded_at": _state["hierarchy_uploaded_at"],
        "row_count": _state["hierarchy_row_count"],
    })


# ---------------------------------------------------------------------------
# Vertical filter (e.g. Inhouse / TC-Channel / Emerging)
# ---------------------------------------------------------------------------
VERTICAL_FIELD = "vertical_id"
VERTICAL_LABELS = {"1": "Inhouse", "56": "TC-Channel", "42": "Emerging"}


def get_available_verticals(rows):
    """Look at whatever's actually in the pushed data rather than hardcoding
    just 1/42/56 — if a new vertical gets added to the query later, it shows
    up here automatically."""
    seen = set()
    for r in rows:
        v = r.get(VERTICAL_FIELD)
        if v is not None and str(v).strip() != "":
            seen.add(normalize_id(v))
    ordered = sorted(seen, key=lambda x: (x not in VERTICAL_LABELS, x))
    return [{"value": v, "label": VERTICAL_LABELS.get(v, f"Vertical {v}")} for v in ordered]


def filter_rows_by_vertical(rows, vertical):
    if not vertical or vertical == "all":
        return rows
    return [r for r in rows if normalize_id(r.get(VERTICAL_FIELD)) == normalize_id(vertical)]


# ---------------------------------------------------------------------------
# Merge + serve
# ---------------------------------------------------------------------------
def get_computed(vertical="all"):
    with _lock:
        if _state["redash_rows"] is None:
            return None
        cache = _state.setdefault("computed_by_vertical", {})
        key = vertical or "all"
        if key not in cache:
            hmap = _state["hierarchy_map"] or {}
            rows = filter_rows_by_vertical(_state["redash_rows"], key)
            computed = merge_and_build_tree(rows, hmap)
            # Authoritative location sales has no vertical breakdown, so this
            # total always reflects ALL verticals combined -- even when a
            # specific vertical is selected here. That's intentional: some
            # sales are attributed directly to L3 (branch managers), which
            # never appear in the L1-sourced employee breakdown at all, so
            # the breakdown's sum and this top-level total are allowed to
            # disagree by design.
            computed["ambiguous_locations"] = []
            computed["unmatched_locations"] = []
            if _state["location_sales_rows"]:
                overlay_result = apply_authoritative_location_sales(
                    computed["tree"], _state["location_sales_rows"]
                )
                computed["ambiguous_locations"] = overlay_result["ambiguous"]
                computed["unmatched_locations"] = overlay_result["unmatched"]
            cache[key] = computed
        return cache[key]


def gzip_json_response(obj):
    payload = json.dumps(obj)
    accept_encoding = request.headers.get("Accept-Encoding", "")
    if "gzip" in accept_encoding:
        compressed = gzip.compress(payload.encode("utf-8"))
        resp = Response(compressed, mimetype="application/json")
        resp.headers["Content-Encoding"] = "gzip"
        resp.headers["Vary"] = "Accept-Encoding"
        return resp
    return Response(payload, mimetype="application/json")


@app.route("/api/dashboard-data")
def api_dashboard_data():
    try:
        vertical = request.args.get("vertical", "all")
        computed = get_computed(vertical=vertical)
        if computed is None:
            return jsonify({
                "ok": False,
                "error": "No data has been pushed yet. Run push_local.py from a machine with Redash access first.",
            })
        payload = tree_payload(computed)
        payload["ok"] = True
        payload["title"] = DASHBOARD_TITLE
        payload["redash_cached_at"] = _state["redash_pushed_at"]
        payload["hierarchy_uploaded_at"] = _state["hierarchy_uploaded_at"]
        payload["hierarchy_row_count"] = _state["hierarchy_row_count"]
        payload["location_sales_pushed_at"] = _state["location_sales_pushed_at"]
        payload["location_sales_row_count"] = len(_state["location_sales_rows"] or [])
        payload["ambiguous_locations"] = computed.get("ambiguous_locations", [])
        payload["unmatched_locations"] = computed.get("unmatched_locations", [])
        payload["served_at"] = time.time()
        payload["selected_vertical"] = vertical
        payload["available_verticals"] = get_available_verticals(_state["redash_rows"])
        return gzip_json_response(payload)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/employee/<employee_id>")
def api_employee(employee_id):
    try:
        vertical = request.args.get("vertical", "all")
        computed = get_computed(vertical=vertical)
        if computed is None:
            return jsonify({"ok": False, "error": "No data has been pushed yet."}), 404
        eid = normalize_id(employee_id)
        row = computed["employees"].get(eid)
        if row is None:
            return jsonify({"ok": False, "error": "Employee not found in latest data"}), 404
        hmap = _state["hierarchy_map"] or {}
        h = hmap.get(eid, {})
        enriched = dict(row)
        enriched["_am"] = h.get("am")
        enriched["_rm"] = h.get("rm")
        enriched["_zm"] = h.get("zm")
        return jsonify({"ok": True, "employee": enriched})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
